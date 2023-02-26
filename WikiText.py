
import wikipedia as w
import openpyxl as o
import requests
from bs4 import BeautifulSoup

w.set_lang('ru')

doc = o.reader.excel.load_workbook(filename="words_temp.xlsx")
doc.active = 7
sheet = doc.active




i=5

for row in sheet.iter_rows(min_row=i, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
    current_row = []
    links = []
    for cell in row:
        current_row.append(cell.value)
    if current_row[3] is not None:
        try:
            search = w.page(current_row[3])
            fnd=search.content
            #выделяем первый абзац
            abz=fnd[:fnd.find("\n")]
            #если в абзаце больше 30 слов, то обрезаем до 30
            if len(abz.split(" "))>30:
                s=abz.split(" ")
                s=s[:30]
                s=' '.join(s)+"..."
                c3 = sheet.cell(row = i, column = 7)
                c3.value = s
                #print(s)
            else:
                c3 = sheet.cell(row = i, column = 7)
                c3.value = abz
                #print(abz)           
        except:
            print("Error"+str(i))
            print(fnd)
            i=i+1
            continue          
    i=i+1
    print(i)

doc.save('words_temp.xlsx')
