#Выкачка начальных страниц по навыкам

import wikipedia as w
import openpyxl as o
import requests
from bs4 import BeautifulSoup

w.set_lang('ru')

doc = o.reader.excel.load_workbook(filename="words_temp.xlsx")
doc.active = 2
sheet = doc.active

i=2

for row in sheet.iter_rows(min_row=i, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
    current_row = []
    for cell in row:
        current_row.append(cell.value)
    if current_row[2] is None:
        skill = current_row[1]
        query="https://ru.wikipedia.org/wiki/"+skill
        print(query)
        try:
            page = requests.get(query)
            soup = BeautifulSoup(page.content, 'html.parser')
            title = soup.find(id="firstHeading").get_text()
            text = soup.find(id="mw-content-text").get_text()
            if "В Википедии нет статьи с таким названием" in text:
                i=i+1
                continue
            else:
                #page = w.page(search[0])
                c4 = sheet.cell(row = i, column = 4)
                c4.value = title
                print(title)
                if title.find(" ") !=-1:
                    l = title.split()
                    title = '_'.join(l)
                query="https://ru.wikipedia.org/wiki/"+title                   
                c3 = sheet.cell(row = i, column = 3)
                c3.value = query
        except:
            print("ErrorName")
            i=i+1
            continue
    i=i+1
        

doc.save('words_temp.xlsx')
