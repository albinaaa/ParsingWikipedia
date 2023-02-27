#Выкачка ссылок на упоминаемые страницы

import wikipedia as w
import openpyxl as o
import requests
from bs4 import BeautifulSoup

w.set_lang('ru')

doc = o.reader.excel.load_workbook(filename="words_temp.xlsx")
doc.active = 2
sheet = doc.active


sheet2 = doc["Темы_Уровень_02b"]

i=2
j=2
for row in sheet.iter_rows(min_row=i, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
    current_row = []
    links = []
    for cell in row:
        current_row.append(cell.value)
    if current_row[3] is not None:
        try:
            search = w.page(current_row[3])
            links = search.links
            for l in links:
                try:
                    query="https://ru.wikipedia.org/wiki/"+l
                    #print(query)
                    page = requests.get(query)
                    soup = BeautifulSoup(page.content, 'html.parser')
                    title = soup.find(id="firstHeading").get_text()
                    text = soup.find(id="mw-content-text").get_text()
                    if "В Википедии нет статьи с таким названием" in text:
                        continue
                    else:                       
                        c1 = sheet2.cell(row = j, column = 1)
                        c1.value = current_row[0]
                        c2 = sheet2.cell(row = j, column = 2)
                        c2.value = current_row[1]
                        c3 = sheet2.cell(row = j, column = 3)
                        c3.value = current_row[2]
                        c4 = sheet2.cell(row = j, column = 4)
                        c4.value = current_row[3]
                        c6 = sheet2.cell(row = j, column = 6)
                        c6.value = title
                        if title.find(" ") !=-1:
                            l = title.split()
                            title = '_'.join(l)
                        url="https://ru.wikipedia.org/wiki/"+title
                        c5 = sheet2.cell(row = j, column = 5)
                        c5.value = url
                        j = j+1                    
                except:
                    #print("ErrorName")
                    continue
        except:
            c1 = sheet2.cell(row = j, column = 1)
            c1.value = current_row[0]
            c2 = sheet2.cell(row = j, column = 2)
            c2.value = current_row[1]
            c3 = sheet2.cell(row = j, column = 3)
            c3.value = current_row[2]
            c4 = sheet2.cell(row = j, column = 4)
            c4.value = current_row[3]
            c5 = sheet2.cell(row = j, column = 5)
            c5.value = "Error"
            continue          
    i=i+1
    #print(i)

doc.save('words_temp.xlsx')
